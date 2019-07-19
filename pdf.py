import os, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class Pdf:
    def __init__(self):
        # Open the workbook
        self.wb = Workbook()
        # Grab the active wroksheet
        self.ws = self.wb.active

        self.header_font = Font(size=14, bold=True)
        self.my_fill = PatternFill(
                        start_color='00008B',
                        end_color='00008B',
                        fill_type='solid'
        ) # Blue background cell

        self.file_name = datetime.datetime.now().strftime("%d-%b")

    def get_pdf(self, querys):
        self.ws.append(["FTML DYEING SECTION",])
        self.ws['A1'].style = 'Title'
        # 2 row empty
        self.ws.append([])
        self.ws.append([])

        # Header
        self.ws.append(['Serial No.', 'Machine No.', 'Send Time', ' Receive Time', 'Status', 'Date'])

        # Give style for header
        for cell in self.ws["4:4"]:
            cell.font = self.header_font
            cell.fill = self.my_fill
        
        # data insert into the rows
        for query in querys:
            self.ws.append(query)

        self.wb.save(os.getcwd()+"/file/"+"{}.xlsx".format(self.file_name))
